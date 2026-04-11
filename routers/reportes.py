from fastapi import APIRouter, Depends
from fastapi.responses import Response
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

from supabase_db import get_sb
from routers.autenticacion import get_current_user

router = APIRouter()

# ── PALETTE ──────────────────────────────────────────────────────────────────
C_AZUL      = "1A4FD6"
C_AZUL_OSC  = "0F2554"
C_TEAL      = "0891B2"
C_VERDE     = "059669"
C_ORO       = "D97706"
C_ROJO      = "DC2626"
C_ROJO_BG   = "FEF2F2"
C_VERDE_BG  = "ECFDF5"
C_ORO_BG    = "FFFBEB"
C_AZUL_BG   = "EEF3FF"
C_GRIS      = "F4F6FA"
C_GRIS_MED  = "E4E9F0"
C_BLANCO    = "FFFFFF"
C_TEXTO     = "0D1B2E"

FMT_COP  = '#,##0.00" COP"'
FMT_CANT = '#,##0'
FMT_DATE = 'DD/MM/YYYY'
FMT_PCT  = '0.00%'

# ── HELPERS ──────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_TEXTO, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _border_thin() -> Border:
    s = Side(style="thin", color=C_GRIS_MED)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color=C_GRIS_MED) -> Border:
    return Border(bottom=Side(style="thin", color=color))

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _header_row(ws, row, cols, fill_color=C_AZUL_OSC, font_color=C_BLANCO,
                height=22):
    """Write a styled header row."""
    ws.row_dimensions[row].height = height
    for c, (col, text) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill = _fill(fill_color)
        cell.font = _font(bold=True, color=font_color, size=10)
        cell.alignment = _align("center")
        cell.border = _border_thin()

def _company_header(ws, nombre, nit, titulo, periodo=""):
    """Insert a 3-row company header block."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"FINPRO CAPITAL  ·  {nombre.upper()}"
    c.fill = _fill(C_AZUL_OSC)
    c.font = _font(bold=True, color=C_BLANCO, size=13)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = f"NIT: {nit}  ·  {titulo}"
    c2.fill = _fill(C_AZUL)
    c2.font = _font(bold=False, color=C_BLANCO, size=10)
    c2.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:H3")
    c3 = ws["A3"]
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    c3.value = f"Generado el {now}  {('· Período: ' + periodo) if periodo else ''}"
    c3.fill = _fill(C_GRIS)
    c3.font = _font(italic=True, color="5A6A7E", size=9)
    c3.alignment = _align("center")
    ws.row_dimensions[3].height = 16

def _totals_row(ws, row, label_col, label, sum_cols, fmt=FMT_COP):
    """Write a totals row with bold styling."""
    ws.row_dimensions[row].height = 20
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(C_AZUL_BG)
        cell.border = Border(
            top=Side(style="medium", color=C_AZUL),
            bottom=Side(style="thin", color=C_GRIS_MED),
        )
    lbl = ws.cell(row=row, column=label_col, value=label)
    lbl.font = _font(bold=True, color=C_AZUL_OSC, size=10)
    lbl.alignment = _align("right")
    for col in sum_cols:
        c = ws.cell(row=row, column=col)
        c.font = _font(bold=True, color=C_AZUL_OSC, size=11)
        c.number_format = fmt
        c.alignment = _align("right")

def _excel_response(wb: openpyxl.Workbook, filename: str) -> Response:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _get_user_data(sb, user_id, nit):
    r = sb.table("usuarios").select("nombre,nit").eq("id", user_id).execute()
    if r.data:
        return r.data[0]["nombre"], r.data[0]["nit"]
    return "EMPRESA", nit


# ── 1. LIBRO DIARIO ──────────────────────────────────────────────────────────
@router.get("/libro-diario", summary="Exportar Libro Diario en Excel")
async def reporte_libro_diario(current_user: dict = Depends(get_current_user)):
    sb = get_sb()
    nombre, nit = _get_user_data(sb, current_user["id"], current_user["nit"])

    # Fetch cesiones
    fids = sb.table("facturas").select("id").eq("emisor_id", current_user["id"]).execute()
    ids = [f["id"] for f in (fids.data or [])]
    ces = []
    if ids:
        r = sb.table("cesiones").select("*").in_("factura_id", ids)\
              .eq("estado", "ACEPTADA").order("fecha_cesion").execute()
        ces = r.data or []

    wb = openpyxl.Workbook()
    wb.properties.creator = "FINPRO CAPITAL"

    # ── Hoja 1: Libro Diario ──
    ws = wb.active
    ws.title = "Libro Diario"
    ws.sheet_view.showGridLines = False

    _company_header(ws, nombre, nit, "LIBRO DIARIO — PUC COLOMBIA")

    COLS = [
        ("A", "FECHA"),
        ("B", "CÓD. PUC"),
        ("C", "CUENTA CONTABLE"),
        ("D", "NÚM. CESIÓN"),
        ("E", "CESIONARIO"),
        ("F", "DÉBITO (COP)"),
        ("G", "CRÉDITO (COP)"),
        ("H", "ESTADO"),
    ]
    _header_row(ws, 5, COLS)

    widths = [14, 10, 42, 18, 28, 20, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    PUC = {
        "1305": "Deudores Comerciales — Clientes Nacionales",
        "4135": "Ingresos Financieros — Cesión de Cartera",
    }

    row = 6
    total_d = 0.0
    total_c = 0.0

    for c in ces:
        val = float(c.get("valor_cesion", 0) or 0)
        fecha = c.get("fecha_cesion", "")[:10]
        num = c.get("numero_cesion", "—")
        ces_nombre = c.get("cesionario_nombre", "—")

        for codigo, es_debito in [("1305", True), ("4135", False)]:
            ws.row_dimensions[row].height = 17
            cuenta = ws.cell(row=row, column=1, value=fecha)
            cuenta.number_format = FMT_DATE
            cuenta.font = _font(size=9)
            cuenta.alignment = _align()
            cuenta.border = _border_thin()

            b = ws.cell(row=row, column=2, value=codigo)
            b.font = _font(bold=True, color=C_AZUL, size=9)
            b.alignment = _align("center")
            b.border = _border_thin()

            cc = ws.cell(row=row, column=3, value=PUC[codigo])
            cc.font = _font(size=9)
            if not es_debito:
                cc.font = _font(size=9, italic=True, color="5A6A7E")
                cc.alignment = _align()
                # indent credit row visually
                cc.value = "    " + PUC[codigo]
            cc.border = _border_thin()

            nc = ws.cell(row=row, column=4, value=num)
            nc.font = _font(size=9, color="5A6A7E")
            nc.alignment = _align("center")
            nc.border = _border_thin()

            ces_c = ws.cell(row=row, column=5, value=ces_nombre)
            ces_c.font = _font(size=9)
            ces_c.border = _border_thin()

            deb = ws.cell(row=row, column=6, value=val if es_debito else None)
            deb.number_format = FMT_COP
            deb.font = _font(bold=es_debito, color=C_TEXTO if es_debito else C_BLANCO, size=9)
            deb.alignment = _align("right")
            deb.border = _border_thin()

            cre = ws.cell(row=row, column=7, value=val if not es_debito else None)
            cre.number_format = FMT_COP
            cre.font = _font(bold=not es_debito, color="5A6A7E" if not es_debito else C_BLANCO, size=9)
            cre.alignment = _align("right")
            cre.border = _border_thin()

            est = ws.cell(row=row, column=8, value="ACEPTADA")
            est.fill = _fill(C_VERDE_BG)
            est.font = _font(bold=True, color=C_VERDE, size=8)
            est.alignment = _align("center")
            est.border = _border_thin()

            if es_debito:
                total_d += val
            else:
                total_c += val
            row += 1

        # Separator between entries
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = Border(
                bottom=Side(style="hair", color=C_GRIS_MED))
        row += 1

    # Totals
    _totals_row(ws, row, 3, "TOTALES DEL PERÍODO", [6, 7])
    ws.cell(row=row, column=6).value = total_d
    ws.cell(row=row, column=7).value = total_c

    # Balance check
    row += 2
    bal = ws.cell(row=row, column=3,
                  value="✓ BALANCE CUADRADO" if abs(total_d - total_c) < 0.01
                  else f"⚠ DIFERENCIA: ${total_d - total_c:,.2f}")
    bal.font = _font(bold=True,
                     color=C_VERDE if abs(total_d - total_c) < 0.01 else C_ROJO,
                     size=11)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:H{row - 3}"

    # ── Hoja 2: Resumen PUC ──
    ws2 = wb.create_sheet("Resumen por Cuenta")
    ws2.sheet_view.showGridLines = False
    _company_header(ws2, nombre, nit, "SALDOS POR CUENTA PUC")
    _header_row(ws2, 5, [("A","CÓD. PUC"),("B","CUENTA"),("C","SALDO DÉBITO"),("D","SALDO CRÉDITO"),("E","NATURALEZA")])
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 44
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 22
    ws2.column_dimensions["E"].width = 14
    data_puc = [
        ("1305", PUC["1305"], total_d, 0, "DEUDORA"),
        ("4135", PUC["4135"], 0, total_c, "ACREEDORA"),
    ]
    for r2, (cod, cta, deb2, cre2, nat) in enumerate(data_puc, start=6):
        ws2.row_dimensions[r2].height = 18
        for ci, val2 in enumerate([cod, cta, deb2, cre2, nat], start=1):
            cell = ws2.cell(row=r2, column=ci, value=val2)
            cell.border = _border_thin()
            cell.font = _font(size=10)
            if ci in (3, 4):
                cell.number_format = FMT_COP
                cell.alignment = _align("right")

    fname = f"libro_diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, fname)


# ── 2. CARTERA ───────────────────────────────────────────────────────────────
@router.get("/cartera", summary="Exportar Análisis de Cartera en Excel")
async def reporte_cartera(current_user: dict = Depends(get_current_user)):
    sb = get_sb()
    nombre, nit = _get_user_data(sb, current_user["id"], current_user["nit"])

    r = sb.table("facturas").select("*").eq("emisor_id", current_user["id"]).execute()
    fs = [f for f in (r.data or []) if f.get("estado") not in ("PAGADA",)]

    hoy = date.today()
    buckets = [
        (None, 0,  "Al Día",        C_VERDE,    C_VERDE_BG),
        (1,   30,  "1 – 30 días",   C_ORO,      C_ORO_BG),
        (31,  60,  "31 – 60 días",  "F97316",   "FFF7ED"),
        (61,  90,  "61 – 90 días",  C_ROJO,     C_ROJO_BG),
        (91, None, "+90 días",      "7F1D1D",   "FEE2E2"),
    ]

    def get_bucket(dias):
        if dias <= 0: return 0
        if dias <= 30: return 1
        if dias <= 60: return 2
        if dias <= 90: return 3
        return 4

    facturas_con_dias = []
    for f in fs:
        try:
            venc = date.fromisoformat(str(f["fecha_vencimiento"])[:10])
            dias = (hoy - venc).days
        except Exception:
            dias = 0
        facturas_con_dias.append({**f, "_dias": dias, "_bucket": get_bucket(dias)})

    wb = openpyxl.Workbook()
    wb.properties.creator = "FINPRO CAPITAL"

    # ── Hoja 1: Resumen Aging ──
    ws = wb.active
    ws.title = "Resumen Cartera"
    ws.sheet_view.showGridLines = False
    _company_header(ws, nombre, nit, "ANÁLISIS DE ANTIGÜEDAD DE CARTERA")

    widths_r = [20, 14, 22, 22, 12]
    for i, w in enumerate(widths_r, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _header_row(ws, 5, [
        ("A","RANGO"), ("B","# FACTURAS"), ("C","VALOR TOTAL (COP)"),
        ("D","% DEL TOTAL"), ("E","SEMÁFORO")
    ])

    total_gral = sum(f.get("valor_total", 0) or 0 for f in fs)
    row = 6
    for bi, (_, _, label, color, bg) in enumerate(buckets):
        items = [f for f in facturas_con_dias if f["_bucket"] == bi]
        total_b = sum(f.get("valor_total", 0) or 0 for f in items)
        pct = total_b / total_gral if total_gral else 0
        ws.row_dimensions[row].height = 20

        for ci, (val2, fmt2, align2) in enumerate([
            (label, "@", "left"),
            (len(items), FMT_CANT, "center"),
            (total_b, FMT_COP, "right"),
            (pct, FMT_PCT, "center"),
            ("●", "@", "center"),
        ], start=1):
            cell = ws.cell(row=row, column=ci, value=val2)
            cell.fill = _fill(bg)
            cell.font = _font(bold=(ci == 1), color=color if ci in (1, 5) else C_TEXTO, size=10)
            cell.number_format = fmt2
            cell.alignment = _align(align2)
            cell.border = _border_thin()
        row += 1

    # Grand total
    _totals_row(ws, row, 1, "TOTAL CARTERA", [3])
    ws.cell(row=row, column=2).value = len(fs)
    ws.cell(row=row, column=2).font = _font(bold=True, size=11)
    ws.cell(row=row, column=2).alignment = _align("center")
    ws.cell(row=row, column=3).value = total_gral

    # Chart (bar)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Distribución de Cartera por Antigüedad"
    chart.y_axis.title = "COP"
    chart.x_axis.title = "Rango"
    chart.style = 10
    chart.height = 12
    chart.width = 20
    data_ref = Reference(ws, min_col=3, max_col=3, min_row=5, max_row=5 + len(buckets))
    cats = Reference(ws, min_col=1, min_row=6, max_row=5 + len(buckets))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "G5")

    ws.freeze_panes = "A6"

    # ── Hoja 2: Detalle ──
    ws2 = wb.create_sheet("Detalle Facturas")
    ws2.sheet_view.showGridLines = False
    _company_header(ws2, nombre, nit, "DETALLE DE FACTURAS EN CARTERA")
    _header_row(ws2, 5, [
        ("A","NÚMERO"), ("B","CUFE"), ("C","ADQUIRIENTE"), ("D","NIT ADQUIRIENTE"),
        ("E","VALOR BASE"), ("F","IVA"), ("G","VALOR TOTAL"), ("H","FECHA EMISION"),
        ("I","FECHA VENC."), ("J","DÍAS MORA"), ("K","RANGO"), ("L","ESTADO"),
    ], fill_color=C_AZUL)
    widths2 = [14, 26, 30, 16, 18, 14, 18, 14, 14, 11, 14, 14]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    facturas_con_dias.sort(key=lambda x: x["_dias"], reverse=True)
    for r2, f in enumerate(facturas_con_dias, start=6):
        bi = f["_bucket"]
        color_b, bg_b = buckets[bi][3], buckets[bi][4]
        ws2.row_dimensions[r2].height = 16
        vals = [
            f"{f.get('prefijo','FV')}-{f.get('numero','')}",
            (f.get("cufe","") or "")[:24]+"…",
            f.get("adquiriente_nombre","—"),
            f.get("adquiriente_nit","—"),
            float(f.get("valor_base",0) or 0),
            float(f.get("valor_iva",0) or 0),
            float(f.get("valor_total",0) or 0),
            str(f.get("fecha_emision",""))[:10],
            str(f.get("fecha_vencimiento",""))[:10],
            f["_dias"] if f["_dias"] > 0 else 0,
            buckets[bi][2],
            f.get("estado","—"),
        ]
        fmts = ["@","@","@","@",FMT_COP,FMT_COP,FMT_COP,FMT_DATE,FMT_DATE,FMT_CANT,"@","@"]
        for ci, (val2, fmt2) in enumerate(zip(vals, fmts), start=1):
            cell = ws2.cell(row=r2, column=ci, value=val2)
            cell.font = _font(size=9)
            cell.number_format = fmt2
            cell.border = _border_thin()
            cell.alignment = _align("right" if ci in (5,6,7,10) else "left")
            if ci == 10 and f["_dias"] > 0:
                cell.fill = _fill(bg_b)
                cell.font = _font(bold=True, color=color_b, size=9)
            if ci == 11:
                cell.fill = _fill(bg_b)
                cell.font = _font(bold=True, color=color_b, size=9)
                cell.alignment = _align("center")

    ws2.auto_filter.ref = f"A5:L{5+len(facturas_con_dias)}"
    ws2.freeze_panes = "A6"

    fname = f"cartera_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, fname)


# ── 3. ESTADO DE RESULTADOS ──────────────────────────────────────────────────
@router.get("/estado-resultados", summary="Exportar Estado de Resultados en Excel")
async def reporte_estado_resultados(current_user: dict = Depends(get_current_user)):
    sb = get_sb()
    nombre, nit = _get_user_data(sb, current_user["id"], current_user["nit"])

    fids = sb.table("facturas").select("id").eq("emisor_id", current_user["id"]).execute()
    ids = [f["id"] for f in (fids.data or [])]
    ces = []
    if ids:
        r = sb.table("cesiones").select("*").in_("factura_id", ids)\
              .eq("estado", "ACEPTADA").order("fecha_cesion").execute()
        ces = r.data or []

    total_ingresos = sum(float(c.get("valor_cesion", 0) or 0) for c in ces)
    gastos_financieros = total_ingresos * 0.005
    utilidad = total_ingresos - gastos_financieros

    # Monthly breakdown
    by_month: dict = {}
    for c in ces:
        key = str(c.get("fecha_cesion",""))[:7]
        if not key: continue
        by_month.setdefault(key, {"cesiones": 0, "valor": 0.0})
        by_month[key]["cesiones"] += 1
        by_month[key]["valor"] += float(c.get("valor_cesion", 0) or 0)
    meses = sorted(by_month.items())

    wb = openpyxl.Workbook()
    wb.properties.creator = "FINPRO CAPITAL"

    # ── Hoja 1: P&L ──
    ws = wb.active
    ws.title = "Estado de Resultados"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24

    _company_header(ws, nombre, nit, "ESTADO DE RESULTADOS")
    ws["A1"].value = f"FINPRO CAPITAL  ·  {nombre.upper()}"
    ws.merge_cells("A1:E1")
    ws.merge_cells("A2:E2")
    ws.merge_cells("A3:E3")

    def pnl_seccion(row, titulo):
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=titulo)
        c.fill = _fill(C_AZUL_OSC)
        c.font = _font(bold=True, color=C_BLANCO, size=11)
        c.alignment = _align("left")
        return row + 1

    def pnl_linea(row, cod, desc, valor, color=C_TEXTO, bold=False, bg=None):
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=2, value=cod).font = _font(color=C_AZUL, size=9, bold=True)
        ws.cell(row=row, column=2).alignment = _align("center")
        ws.cell(row=row, column=3, value=desc).font = _font(size=10, bold=bold)
        ws.cell(row=row, column=3).border = _border_thin()
        cell_v = ws.cell(row=row, column=4, value=valor)
        cell_v.number_format = FMT_COP
        cell_v.font = _font(bold=bold, color=color, size=10)
        cell_v.alignment = _align("right")
        cell_v.border = _border_thin()
        if bg:
            for ci in range(1, 6):
                ws.cell(row=row, column=ci).fill = _fill(bg)
        for ci in [2, 3, 4]:
            ws.cell(row=row, column=ci).border = _border_thin()
        return row + 1

    def pnl_total(row, desc, valor, color=C_AZUL_OSC):
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"B{row}:C{row}")
        c = ws.cell(row=row, column=2, value=desc)
        c.font = _font(bold=True, color=C_BLANCO, size=11)
        c.fill = _fill(C_AZUL)
        c.alignment = _align("right")
        cv = ws.cell(row=row, column=4, value=valor)
        cv.number_format = FMT_COP
        cv.font = _font(bold=True, color=C_BLANCO, size=12)
        cv.fill = _fill(C_AZUL)
        cv.alignment = _align("right")
        return row + 2

    row = 5
    row = pnl_seccion(row, "  INGRESOS OPERACIONALES")
    row = pnl_linea(row, "4135", "Ingresos por cesión de cartera", total_ingresos,
                    color=C_VERDE, bg=C_VERDE_BG)
    row = pnl_total(row, "TOTAL INGRESOS", total_ingresos, C_VERDE)

    row = pnl_seccion(row, "  GASTOS OPERACIONALES")
    row = pnl_linea(row, "5105", "Gastos financieros DIAN (0.5%)", gastos_financieros,
                    color=C_ROJO, bg=C_ROJO_BG)
    row = pnl_total(row, "TOTAL GASTOS", gastos_financieros, C_ROJO)

    row = pnl_seccion(row, "  RESULTADO DEL PERÍODO")
    row = pnl_linea(row, "3305", "Utilidad neta del ejercicio", utilidad,
                    color=C_VERDE if utilidad >= 0 else C_ROJO, bold=True,
                    bg=C_VERDE_BG if utilidad >= 0 else C_ROJO_BG)

    # KPIs box
    row += 2
    for ci, (lbl, val2, fmt2) in enumerate([
        ("# Cesiones aceptadas", len(ces), FMT_CANT),
        ("Valor promedio / cesión", total_ingresos / max(len(ces), 1), FMT_COP),
        ("Margen operacional", (utilidad / max(total_ingresos, 1)), FMT_PCT),
    ], start=2):
        ws.cell(row=row, column=ci, value=lbl).font = _font(bold=True, color=C_AZUL, size=9)
        ws.cell(row=row, column=ci).fill = _fill(C_AZUL_BG)
        ws.cell(row=row, column=ci).border = _border_thin()
        ws.cell(row=row, column=ci).alignment = _align("center")
        ws.cell(row=row+1, column=ci, value=val2).number_format = fmt2
        ws.cell(row=row+1, column=ci).font = _font(bold=True, size=13, color=C_TEXTO)
        ws.cell(row=row+1, column=ci).border = _border_thin()
        ws.cell(row=row+1, column=ci).alignment = _align("center")

    # ── Hoja 2: Detalle por mes ──
    ws2 = wb.create_sheet("Por Mes")
    ws2.sheet_view.showGridLines = False
    _company_header(ws2, nombre, nit, "INGRESOS POR CESIÓN — DETALLE MENSUAL")
    _header_row(ws2, 5, [("A","MES"),("B","# CESIONES"),("C","VALOR TOTAL"),("D","PROMEDIO"),("E","% DEL TOTAL")])
    for i, w in enumerate([16, 14, 22, 22, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r2, (mes, data) in enumerate(meses, start=6):
        ws2.row_dimensions[r2].height = 17
        pct2 = data["valor"] / total_ingresos if total_ingresos else 0
        vals2 = [mes, data["cesiones"], data["valor"],
                 data["valor"]/max(data["cesiones"],1), pct2]
        fmts2 = ["@", FMT_CANT, FMT_COP, FMT_COP, FMT_PCT]
        for ci2, (v2, f2) in enumerate(zip(vals2, fmts2), start=1):
            cell = ws2.cell(row=r2, column=ci2, value=v2)
            cell.number_format = f2
            cell.font = _font(size=10)
            cell.border = _border_thin()
            cell.alignment = _align("right" if ci2 > 1 else "left")

    _totals_row(ws2, 6 + len(meses), 1, "TOTAL", [3, 4])
    ws2.cell(row=6+len(meses), column=2).value = len(ces)
    ws2.cell(row=6+len(meses), column=3).value = total_ingresos
    ws2.cell(row=6+len(meses), column=4).value = total_ingresos / max(len(ces), 1)

    # ── Hoja 3: Detalle cesiones ──
    ws3 = wb.create_sheet("Detalle Cesiones")
    ws3.sheet_view.showGridLines = False
    _company_header(ws3, nombre, nit, "DETALLE DE CESIONES ACEPTADAS")
    _header_row(ws3, 5, [
        ("A","FECHA"), ("B","NÚM. CESIÓN"), ("C","CEDENTE"), ("D","CESIONARIO"),
        ("E","VALOR CESIÓN"), ("F","CUDE"), ("G","ESTADO"),
    ], fill_color=C_TEAL)
    for i, w in enumerate([14, 18, 28, 28, 20, 24, 12], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    for r3, c in enumerate(ces, start=6):
        ws3.row_dimensions[r3].height = 16
        for ci3, val3 in enumerate([
            str(c.get("fecha_cesion",""))[:10],
            c.get("numero_cesion","—"),
            c.get("cedente_nombre","—"),
            c.get("cesionario_nombre","—"),
            float(c.get("valor_cesion",0) or 0),
            (c.get("cude","") or "")[:20]+"…",
            c.get("estado","—"),
        ], start=1):
            cell = ws3.cell(row=r3, column=ci3, value=val3)
            cell.font = _font(size=9)
            cell.border = _border_thin()
            if ci3 == 5:
                cell.number_format = FMT_COP
                cell.alignment = _align("right")
            if ci3 == 7:
                cell.fill = _fill(C_VERDE_BG)
                cell.font = _font(bold=True, color=C_VERDE, size=9)
                cell.alignment = _align("center")
    ws3.auto_filter.ref = f"A5:G{5+len(ces)}"
    ws3.freeze_panes = "A6"

    fname = f"estado_resultados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, fname)


# ── 4. FLUJO DE CAJA ─────────────────────────────────────────────────────────
@router.get("/flujo-caja", summary="Exportar Flujo de Caja Proyectado en Excel")
async def reporte_flujo_caja(current_user: dict = Depends(get_current_user)):
    sb = get_sb()
    nombre, nit = _get_user_data(sb, current_user["id"], current_user["nit"])

    r = sb.table("facturas").select("*").eq("emisor_id", current_user["id"]).execute()
    fs = [f for f in (r.data or []) if f.get("estado") not in ("PAGADA",)]
    hoy = date.today()

    by_month: dict = {}
    for f in fs:
        try:
            venc = date.fromisoformat(str(f["fecha_vencimiento"])[:10])
        except Exception:
            continue
        key = venc.strftime("%Y-%m")
        lbl = venc.strftime("%b %Y")
        by_month.setdefault(key, {"label": lbl, "total": 0.0, "count": 0, "items": []})
        by_month[key]["total"] += float(f.get("valor_total", 0) or 0)
        by_month[key]["count"] += 1
        by_month[key]["items"].append(f)
    meses = sorted(by_month.items())

    wb = openpyxl.Workbook()
    wb.properties.creator = "FINPRO CAPITAL"

    # ── Hoja 1: Proyección mensual ──
    ws = wb.active
    ws.title = "Flujo Proyectado"
    ws.sheet_view.showGridLines = False
    _company_header(ws, nombre, nit, "FLUJO DE CAJA PROYECTADO")

    for i, w in enumerate([14, 16, 22, 22, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _header_row(ws, 5, [
        ("A","MES"), ("B","# FACTURAS"), ("C","ENTRADA ESPERADA (COP)"),
        ("D","ACUMULADO (COP)"), ("E","% DEL TOTAL"), ("F","ESTADO"),
    ])

    total_flujo = sum(m["total"] for _, m in meses)
    acum = 0.0
    for row, (key, m) in enumerate(meses, start=6):
        acum += m["total"]
        pct = m["total"] / total_flujo if total_flujo else 0
        mes_date = date.fromisoformat(key + "-01")
        estado = "VENCIDO" if mes_date < hoy.replace(day=1) else \
                 ("ESTE MES" if mes_date.month == hoy.month and mes_date.year == hoy.year
                  else "PROYECTADO")
        color_est = C_ROJO if estado == "VENCIDO" else \
                    (C_ORO if estado == "ESTE MES" else C_VERDE)
        bg_est = C_ROJO_BG if estado == "VENCIDO" else \
                 (C_ORO_BG if estado == "ESTE MES" else C_VERDE_BG)

        ws.row_dimensions[row].height = 18
        for ci, (val2, fmt2, aln) in enumerate([
            (m["label"], "@", "left"),
            (m["count"], FMT_CANT, "center"),
            (m["total"], FMT_COP, "right"),
            (acum, FMT_COP, "right"),
            (pct, FMT_PCT, "center"),
            (estado, "@", "center"),
        ], start=1):
            cell = ws.cell(row=row, column=ci, value=val2)
            cell.number_format = fmt2
            cell.font = _font(size=10)
            cell.border = _border_thin()
            cell.alignment = _align(aln)
            if ci == 6:
                cell.fill = _fill(bg_est)
                cell.font = _font(bold=True, color=color_est, size=9)

    _totals_row(ws, 6 + len(meses), 1, "TOTAL PROYECTADO", [3, 4])
    ws.cell(row=6+len(meses), column=2).value = sum(m["count"] for _, m in meses)
    ws.cell(row=6+len(meses), column=3).value = total_flujo
    ws.cell(row=6+len(meses), column=4).value = total_flujo

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:F{5+len(meses)}"

    # ── Hoja 2: Calendario de vencimientos ──
    ws2 = wb.create_sheet("Calendario Vencimientos")
    ws2.sheet_view.showGridLines = False
    _company_header(ws2, nombre, nit, "CALENDARIO DE VENCIMIENTOS")
    _header_row(ws2, 5, [
        ("A","FACTURA"), ("B","ADQUIRIENTE"), ("C","NIT"),
        ("D","VALOR TOTAL"), ("E","FECHA EMISION"), ("F","FECHA VENC."),
        ("G","DÍAS"), ("H","ESTADO"),
    ], fill_color=C_TEAL)
    for i, w in enumerate([16, 30, 16, 20, 14, 14, 10, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    fs_sorted = sorted(fs, key=lambda f: str(f.get("fecha_vencimiento",""))[:10])
    for r2, f in enumerate(fs_sorted, start=6):
        try:
            venc = date.fromisoformat(str(f["fecha_vencimiento"])[:10])
            dias = (venc - hoy).days
        except Exception:
            dias = 0
            venc = hoy
        color_d = C_ROJO if dias < 0 else (C_ORO if dias <= 30 else C_VERDE)
        bg_d = C_ROJO_BG if dias < 0 else (C_ORO_BG if dias <= 30 else C_VERDE_BG)
        ws2.row_dimensions[r2].height = 16
        for ci, (val2, fmt2, aln) in enumerate([
            (f"{f.get('prefijo','FV')}-{f.get('numero','')}", "@", "left"),
            (f.get("adquiriente_nombre","—"), "@", "left"),
            (f.get("adquiriente_nit","—"), "@", "center"),
            (float(f.get("valor_total",0) or 0), FMT_COP, "right"),
            (str(f.get("fecha_emision",""))[:10], FMT_DATE, "center"),
            (str(f.get("fecha_vencimiento",""))[:10], FMT_DATE, "center"),
            (abs(dias), FMT_CANT, "center"),
            (f.get("estado","—"), "@", "center"),
        ], start=1):
            cell = ws2.cell(row=r2, column=ci, value=val2)
            cell.number_format = fmt2
            cell.font = _font(size=9)
            cell.border = _border_thin()
            cell.alignment = _align(aln)
            if ci == 7:
                cell.fill = _fill(bg_d)
                cell.font = _font(bold=True, color=color_d, size=9)

    ws2.auto_filter.ref = f"A5:H{5+len(fs)}"
    ws2.freeze_panes = "A6"

    fname = f"flujo_caja_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, fname)


# ── 5. REPORTE CONSOLIDADO ───────────────────────────────────────────────────
@router.get("/consolidado", summary="Exportar Reporte Consolidado (todas las hojas)")
async def reporte_consolidado(current_user: dict = Depends(get_current_user)):
    """Genera un Excel con 6 hojas: Dashboard, Libro Diario, Cartera,
    Flujo de Caja, Estado de Resultados, Detalle Cesiones."""
    sb = get_sb()
    nombre, nit = _get_user_data(sb, current_user["id"], current_user["nit"])

    fids = sb.table("facturas").select("id").eq("emisor_id", current_user["id"]).execute()
    ids = [f["id"] for f in (fids.data or [])]
    ces = []
    if ids:
        rc = sb.table("cesiones").select("*").in_("factura_id", ids)\
               .order("fecha_cesion").execute()
        ces = rc.data or []

    rf = sb.table("facturas").select("*").eq("emisor_id", current_user["id"]).execute()
    fs = rf.data or []

    hoy = date.today()
    ces_ok = [c for c in ces if c.get("estado") == "ACEPTADA"]
    total_ingresos = sum(float(c.get("valor_cesion", 0) or 0) for c in ces_ok)
    gastos = total_ingresos * 0.005
    utilidad = total_ingresos - gastos

    wb = openpyxl.Workbook()
    wb.properties.creator = "FINPRO CAPITAL"

    # ── Dashboard ──
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    _company_header(ws, nombre, nit, "REPORTE CONSOLIDADO FINPRO CAPITAL")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 18

    kpis = [
        ("Total Facturas", len(fs), FMT_CANT, C_AZUL_OSC, C_AZUL_BG),
        ("Facturas Cedidas", sum(1 for f in fs if f.get("estado")=="CEDIDA"), FMT_CANT, C_VERDE, C_VERDE_BG),
        ("Cesiones Aceptadas RADIAN", len(ces_ok), FMT_CANT, C_TEAL, "F0FDFF"),
        ("Total Ingresos (COP)", total_ingresos, FMT_COP, C_VERDE, C_VERDE_BG),
        ("Total Gastos (COP)", gastos, FMT_COP, C_ORO, C_ORO_BG),
        ("Utilidad Neta (COP)", utilidad, FMT_COP,
         C_VERDE if utilidad >= 0 else C_ROJO,
         C_VERDE_BG if utilidad >= 0 else C_ROJO_BG),
        ("Cartera Total Vigente",
         sum(float(f.get("valor_total",0) or 0) for f in fs if f.get("estado") not in ("PAGADA","CEDIDA")),
         FMT_COP, "7C3AED", "F5F3FF"),
    ]
    row = 5
    _header_row(ws, row, [("A","INDICADOR"),("B","VALOR"),("C","ESTADO")])
    row += 1
    for lbl, val2, fmt2, color2, bg2 in kpis:
        ws.row_dimensions[row].height = 22
        cl = ws.cell(row=row, column=1, value=lbl)
        cl.fill = _fill(bg2)
        cl.font = _font(bold=True, color=color2, size=10)
        cl.border = _border_thin()

        cv = ws.cell(row=row, column=2, value=val2)
        cv.number_format = fmt2
        cv.font = _font(bold=True, color=color2, size=12)
        cv.fill = _fill(bg2)
        cv.alignment = _align("right")
        cv.border = _border_thin()

        cs = ws.cell(row=row, column=3, value="●")
        cs.fill = _fill(bg2)
        cs.font = _font(bold=True, color=color2, size=14)
        cs.alignment = _align("center")
        cs.border = _border_thin()
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = \
        _font(italic=True, color="9DAEC0", size=9)

    # ── Libro Diario (simplified) ──
    ws2 = wb.create_sheet("Libro Diario")
    ws2.sheet_view.showGridLines = False
    _company_header(ws2, nombre, nit, "LIBRO DIARIO")
    _header_row(ws2, 5, [
        ("A","FECHA"),("B","CÓD PUC"),("C","CUENTA"),
        ("D","REFERENCIA"),("E","DÉBITO"),("F","CRÉDITO"),
    ])
    for i, w in enumerate([14, 10, 44, 20, 22, 22], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    row2 = 6
    td2 = tc2 = 0.0
    PUC = {"1305": "Deudores Comerciales — Clientes Nacionales",
           "4135": "Ingresos Financieros — Cesión de Cartera"}
    for c in ces_ok:
        val = float(c.get("valor_cesion", 0) or 0)
        fecha = str(c.get("fecha_cesion",""))[:10]
        ref = c.get("numero_cesion","—")
        for cod, deb, cre in [("1305", val, 0), ("4135", 0, val)]:
            for ci, v in enumerate([fecha, cod, PUC[cod], ref, deb or None, cre or None], start=1):
                cell = ws2.cell(row=row2, column=ci, value=v)
                cell.font = _font(size=9)
                cell.border = _border_thin()
                if ci in (5, 6):
                    cell.number_format = FMT_COP
                    cell.alignment = _align("right")
            td2 += deb; tc2 += cre
            row2 += 1
    _totals_row(ws2, row2, 3, "TOTALES", [5, 6])
    ws2.cell(row=row2, column=5).value = td2
    ws2.cell(row=row2, column=6).value = tc2
    ws2.freeze_panes = "A6"

    # ── Cartera ──
    ws3 = wb.create_sheet("Cartera")
    ws3.sheet_view.showGridLines = False
    _company_header(ws3, nombre, nit, "ANÁLISIS DE CARTERA")
    _header_row(ws3, 5, [
        ("A","FACTURA"),("B","ADQUIRIENTE"),("C","VALOR TOTAL"),
        ("D","VENCIMIENTO"),("E","DÍAS MORA"),("F","RANGO"),("G","ESTADO"),
    ])
    for i, w in enumerate([16, 30, 20, 14, 11, 14, 14], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    fs_vigentes = [f for f in fs if f.get("estado") not in ("PAGADA",)]
    for r3, f in enumerate(fs_vigentes, start=6):
        try:
            venc = date.fromisoformat(str(f["fecha_vencimiento"])[:10])
            dias = (hoy - venc).days
        except Exception:
            dias = 0
        rango = "Al día" if dias<=0 else ("1-30d" if dias<=30 else
                ("31-60d" if dias<=60 else ("61-90d" if dias<=90 else "+90d")))
        color_r = C_VERDE if dias<=0 else (C_ORO if dias<=30 else
                  ("F97316" if dias<=60 else (C_ROJO if dias<=90 else "7F1D1D")))
        bg_r = C_VERDE_BG if dias<=0 else (C_ORO_BG if dias<=30 else
               ("FFF7ED" if dias<=60 else C_ROJO_BG))
        for ci, (v3, fmt3) in enumerate([
            (f"{f.get('prefijo','FV')}-{f.get('numero','')}", "@"),
            (f.get("adquiriente_nombre","—"), "@"),
            (float(f.get("valor_total",0) or 0), FMT_COP),
            (str(f.get("fecha_vencimiento",""))[:10], FMT_DATE),
            (max(dias, 0), FMT_CANT),
            (rango, "@"),
            (f.get("estado","—"), "@"),
        ], start=1):
            cell = ws3.cell(row=r3, column=ci, value=v3)
            cell.number_format = fmt3
            cell.font = _font(size=9)
            cell.border = _border_thin()
            if ci in (5, 6):
                cell.fill = _fill(bg_r)
                cell.font = _font(bold=True, color=color_r, size=9)
                cell.alignment = _align("center")
    ws3.auto_filter.ref = f"A5:G{5+len(fs_vigentes)}"
    ws3.freeze_panes = "A6"

    # ── Estado de Resultados (summary) ──
    ws4 = wb.create_sheet("Estado de Resultados")
    ws4.sheet_view.showGridLines = False
    _company_header(ws4, nombre, nit, "ESTADO DE RESULTADOS")
    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 44
    ws4.column_dimensions["C"].width = 26
    for r4, (cod4, desc4, val4, color4, bg4) in enumerate([
        ("4135", "Ingresos por cesión de cartera",  total_ingresos, C_VERDE,    C_VERDE_BG),
        ("",     "  TOTAL INGRESOS",                total_ingresos, C_VERDE,    C_VERDE_BG),
        ("5105", "Gastos financieros DIAN (0.5%)",  -gastos,        C_ROJO,     C_ROJO_BG),
        ("",     "  TOTAL GASTOS",                  -gastos,        C_ROJO,     C_ROJO_BG),
        ("3305", "UTILIDAD NETA DEL EJERCICIO",     utilidad,
         C_VERDE if utilidad >= 0 else C_ROJO,
         C_VERDE_BG if utilidad >= 0 else C_ROJO_BG),
    ], start=5):
        is_total = desc4.startswith("  TOTAL") or "UTILIDAD" in desc4
        ws4.row_dimensions[r4].height = 20 if is_total else 17
        ws4.cell(row=r4, column=1, value=cod4).font = _font(bold=True, color=C_AZUL, size=9)
        ws4.cell(row=r4, column=1).border = _border_thin()
        c_desc = ws4.cell(row=r4, column=2, value=desc4)
        c_desc.fill = _fill(bg4)
        c_desc.font = _font(bold=is_total, color=color4, size=10 if is_total else 9)
        c_desc.border = _border_thin()
        c_val = ws4.cell(row=r4, column=3, value=abs(val4))
        c_val.number_format = FMT_COP
        c_val.fill = _fill(bg4)
        c_val.font = _font(bold=is_total, color=color4, size=11 if is_total else 10)
        c_val.alignment = _align("right")
        c_val.border = _border_thin()

    fname = f"consolidado_finpro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return _excel_response(wb, fname)
